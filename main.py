from sympy import diff, sqrt, simplify, S
from statistics import mean
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy.stats import norm


def get_columns(workbook, sheet=0, _col=0):
    """
    reads excel worksheet and returns a list of the columns

    :param workbook: openpyxl workbook object
    :param sheet: (optional) sheet which needs to be read
    :param _col: (optional) min column
    :return: list of columns
    """
    ws = workbook[workbook.sheetnames[sheet]]
    return [col for col in ws.iter_cols(min_col=_col)]


def transform_cols_to_dict(cols):
    """
    transforms list of given columns into usable dictionary for further calculations

    :param cols: list of columns e.g. from excel file
    :return: dictionary, keys represent terms and values are list of values
    """
    dic = {}
    for col in cols:
        key = None
        values = []
        for cell in col:
            cell = cell.value
            if cell is col[0].value:
                if type(cell) is str:
                    key = S(cell[0])
            else:
                if cell:
                    values.append(cell)
        dic.update({key: values})
    return dic


def transform_cols_to_equation(cols):
    """
    transforms given columns, e.g. read from excel file, into sympy friendly multiplications and symbols

    :param cols: list of columns e.g. from excel file
    :return: multiplication and sympy symbols
    """
    eq = ""
    mul = ""
    syms = []
    for col in cols:
        if col is cols[0]:
            if col[1].value is not None:
                eq = col[1].value
            syms.append(S(col[0].value[0]))
        else:
            mul += f'({col[0].value})*'
            syms.append(S(col[0].value[0]))
    res_eq = simplify(eq) if eq else simplify(mul[0:len(mul) - 1])
    return res_eq, syms


def transform_dict_to_mean_dict(s):
    """
    transforms s (see below) into simple dictionary with given term and mean value

    :param s: dictionary; keys represent term names and values represent a list of the mean value of and the standard
              deviation of them
    :return: transformed simpler dictionary
    """
    dic = {}
    for term, t_values in s.items():
        dic.update({term: t_values[0]})
    return dic


def standard_deviation(mean_dict):
    """
    produces a dictionary filled with mean and standard deviation of given term

    :param mean_dict: dictionary filled with term symbol and list of measured terms
    :return: dictionary; keys represent term names and values represent a list of the mean value of and the standard
             deviation of them
    """
    sigma_dict = {}
    for term, t_values in mean_dict.items():
        if len(t_values) == 2:
            sigma_dict.update({term: t_values})
        else:
            t_mean = round(mean(t_values), 3)
            sigma = 0
            for value in t_values:
                sigma += (value - t_mean) ** 2
            sigma = round(sqrt(sigma / (len(t_values) - 1)), 3)
            sigma_dict.update({term: [t_mean, sigma]})
    return sigma_dict


def function_mean(f, s):
    """
    calculates mean value of function

    :param f: sympy Mul object
    :param s: dictionary; keys represent term names and values represent a list of the mean value of and the standard
              deviation of them
    :return: mean value of given function
    """
    dic = transform_dict_to_mean_dict(s)
    return round(f.subs(dic).evalf(), 3)


def gauss_error_propagation(f, s):
    """
    function produces the standard deviation of given function

    :param f: sympy Mul object
    :param s: dictionary; keys represent term names and values represent a list of the mean value of and the standard
              deviation of them
    :return: standard deviation of given function
    """
    res = 0
    dic = transform_dict_to_mean_dict(s)
    for term, t_values in s.items():
        res += (diff(f, term).subs(dic) ** 2) * (t_values[1] ** 2)
    return round(sqrt(res).evalf(), 3)


def plot_gaussian_distribution(func_mean, sigma, function_name, _fig_counter=1):
    """
    plots the gaussian distribution as graph with given parameters

    source: https://moonbooks.org/Articles/How-to-plot-a-normal-distribution-with-matplotlib-in-python-/

    :param func_mean: mean value of function
    :param sigma: standard deviation of function
    :param function_name: description for plot
    :param _fig_counter: describing number for saving figure file
    :return: figure file name
    """
    func_mean = np.float(func_mean)
    sigma = np.float(sigma)
    x_axis = np.linspace(func_mean - (sigma * 3.5),
                         func_mean + (sigma * 3.5), 100)
    y_axis = norm.pdf(x_axis, func_mean, sigma)
    plt.plot(x_axis, y_axis, color='black')
    #####################################
    pt1 = func_mean + sigma
    plt.plot([pt1, pt1], [0.0, norm.pdf(pt1, func_mean, sigma)], color='black')

    pt2 = func_mean - sigma
    plt.plot([pt2, pt2], [0.0, norm.pdf(pt2, func_mean, func_mean)], color='black')

    ptx = np.linspace(pt1, pt2, 10)
    pty = norm.pdf(ptx, func_mean, sigma)

    plt.fill_between(ptx, pty, color='#0b559f', alpha=1.0)
    #######################################
    pt1 = func_mean + sigma
    plt.plot([pt1, pt1], [0.0, norm.pdf(pt1, func_mean, sigma)], color='black')

    pt2 = func_mean + 2.0 * sigma
    plt.plot([pt2, pt2], [0.0, norm.pdf(pt2, func_mean, sigma)], color='black')

    ptx = np.linspace(pt1, pt2, 10)
    pty = norm.pdf(ptx, func_mean, sigma)

    plt.fill_between(ptx, pty, color='#2b7bba', alpha=1.0)
    ########################################
    pt1 = func_mean - sigma
    plt.plot([pt1, pt1], [0.0, norm.pdf(pt1, func_mean, sigma)], color='black')

    pt2 = func_mean - 2.0 * sigma
    plt.plot([pt2, pt2], [0.0, norm.pdf(pt2, func_mean, sigma)], color='black')

    ptx = np.linspace(pt1, pt2, 10)
    pty = norm.pdf(ptx, func_mean, sigma)

    plt.fill_between(ptx, pty, color='#2b7bba', alpha=1.0)
    ##########################################
    pt1 = func_mean + 2.0 * sigma
    plt.plot([pt1, pt1], [0.0, norm.pdf(pt1, func_mean, sigma)], color='black')

    pt2 = func_mean + 3.0 * sigma
    plt.plot([pt2, pt2], [0.0, norm.pdf(pt2, func_mean, sigma)], color='black')

    ptx = np.linspace(pt1, pt2, 10)
    pty = norm.pdf(ptx, func_mean, sigma)

    plt.fill_between(ptx, pty, color='#539ecd', alpha=1.0)
    ##########################################
    pt1 = func_mean - 2.0 * sigma
    plt.plot([pt1, pt1], [0.0, norm.pdf(pt1, func_mean, sigma)], color='black')

    pt2 = func_mean - 3.0 * sigma
    plt.plot([pt2, pt2], [0.0, norm.pdf(pt2, func_mean, sigma)], color='black')

    ptx = np.linspace(pt1, pt2, 10)
    pty = norm.pdf(ptx, func_mean, sigma)

    plt.fill_between(ptx, pty, color='#539ecd', alpha=1.0)
    ##########################################
    plt.xlabel(function_name)
    plt.grid(alpha=0.6)
    plt.ylim(0, norm.pdf(func_mean, func_mean, sigma) * 1.1)

    sigma_one = mpatches.Patch(color='#0b559f', label='68% confidence interval')
    sigma_two = mpatches.Patch(color='#2b7bba', label='95% confidence interval')
    sigma_three = mpatches.Patch(color='#539ecd', label='99.7% confidence interval')
    plt.legend(handles=[sigma_one, sigma_two, sigma_three], loc='upper right')

    file_name = f'gaussian_distribution_plot_figure{_fig_counter}.png'
    plt.savefig(file_name)
    return file_name


if __name__ == '__main__':
    wb = load_workbook('Thermo.xlsx')
    cols = get_columns(wb)
    s = transform_cols_to_dict(cols)
    function, syms = transform_cols_to_equation(cols)
    s.pop(list(s.keys())[0])
    s = standard_deviation(s)
    res_median = function_mean(function, s)
    res_sigma = gauss_error_propagation(function, s)
    print(f'{syms[0]} = {res_median} +/- {res_sigma}')
    plot_gaussian_distribution(res_median, res_sigma, syms[0])
